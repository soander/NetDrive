import os
import hashlib
from openpyxl import load_workbook
from config import settings


def md5_convert(user_name, password):
    md5_obj = hashlib.md5(user_name.encode())
    md5_obj.update(password.encode())
    return md5_obj.hexdigest()


def verify_register(user_name):
    wb = load_workbook(settings.USER_FILE_PATH)  # 打开excel表格
    sheet = wb.worksheets[0]
    # 通过遍历对比表格中已存在信息
    for i in range(sheet.max_row + 2):
        user_value = sheet.cell(i + 1, 1).value
        if user_name == user_value:
            return True
    return False


def verify(user, psd):
    wb = load_workbook(settings.USER_FILE_PATH)  # 打开excel表格
    sheet = wb.worksheets[0]
    # 通过遍历对比表格中已存在信息
    for i in range(sheet.max_row + 2):
        user_value = sheet.cell(i + 1, 1).value
        psd_value = sheet.cell(i + 1, 2).value
        # 比对成功
        if user == user_value and psd == psd_value:
            return True
    # 比对失败
    return False


def update_info(user_name, pass_word, registertime):
    wb = load_workbook(settings.USER_FILE_PATH)
    sheet = wb.worksheets[0]
    sheet.append([user_name, pass_word, registertime])  # 在最后追加用户信息
    wb.save(settings.USER_FILE_PATH)
    download_path = os.path.join(settings.DB_PATH, user_name)  # 上传或下载文件的路径
    if not os.path.exists(download_path):  # 如果不存在这个路径则创建文件夹
        os.mkdir(download_path)


if __name__ == '__main__':
    verify_result = verify('coco', 123456)
    print(verify_result)
